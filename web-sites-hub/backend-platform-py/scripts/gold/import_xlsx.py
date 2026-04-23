"""Import gold Excel files under data/ into SQLite app.db (sheet-level tables).

Usage:
    cd web-sites-hub/backend-platform-py
    python scripts/gold/import_xlsx.py
"""

from __future__ import annotations

import re
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT_DIR / "data"
DB_PATH = DATA_DIR / "app.db"
XLSX_FILES = [
    DATA_DIR / "gold_price_since_1978.xlsx",
    DATA_DIR / "gold-premiums.xlsx",
]


def normalize_identifier(raw: str, fallback: str) -> str:
    text = (raw or "").strip().lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = fallback
    if text[0].isdigit():
        text = f"c_{text}"
    return text


def normalize_table_name(path: Path, sheet_name: str) -> str:
    return normalize_identifier(f"{path.stem}_{sheet_name}", "gold_data")


def to_cell_value(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    s = str(v).strip()
    return s if s else None


def is_date_like(v) -> bool:
    return isinstance(v, (datetime, date))


def is_meta_line(text: str) -> bool:
    t = text.lower()
    return (
        "source:" in t
        or "important information" in t
        or "for more details" in t
        or "gold market premium" in t
        or "methodology" in t
    )


def detect_header(ws, data_row_idx: int, date_col_idx: int) -> list[str]:
    """Detect header near data section; fallback to generated names."""
    candidate = None
    for i in range(max(1, data_row_idx - 6), data_row_idx):
        vals = [to_cell_value(x) for x in next(ws.iter_rows(min_row=i, max_row=i, values_only=True))]
        non_empty = [v for v in vals if v is not None]
        if len(non_empty) < 2:
            continue
        # Skip obvious metadata long text lines.
        if any(isinstance(v, str) and is_meta_line(v) for v in non_empty):
            continue
        candidate = vals
    if candidate is None:
        return ["date", "value"]

    sliced = candidate[date_col_idx:]
    if not sliced:
        return ["date", "value"]
    headers = [str(x or "").strip() for x in sliced]
    # If header line is mostly empty, fallback.
    if sum(1 for h in headers if h) <= 1:
        return ["date", "value"]
    return headers


def ensure_unique_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for i, c in enumerate(columns):
        base = normalize_identifier(c, f"col_{i+1}")
        count = seen.get(base, 0)
        if count == 0:
            result.append(base)
        else:
            result.append(f"{base}_{count+1}")
        seen[base] = count + 1
    return result


def import_one_sheet(conn: sqlite3.Connection, path: Path, sheet_name: str) -> tuple[str, int]:
    wb = load_workbook(path, data_only=False, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    data_start_idx = None
    date_col_idx = None
    for i, row in enumerate(rows, start=1):
        if not row:
            continue
        for j, v in enumerate(row):
            if is_date_like(v):
                data_start_idx = i
                date_col_idx = j
                break
        if data_start_idx is not None:
            break

    if data_start_idx is None or date_col_idx is None:
        return normalize_table_name(path, sheet_name), 0

    wb2 = load_workbook(path, data_only=False, read_only=True)
    ws2 = wb2[sheet_name]
    headers = detect_header(ws2, data_start_idx, date_col_idx)
    wb2.close()
    cols = ensure_unique_columns(headers)
    table = normalize_table_name(path, sheet_name)

    col_defs = ", ".join([f'"{c}" TEXT' for c in cols])
    conn.execute(f'DROP TABLE IF EXISTS "{table}"')
    conn.execute(
        f"""
        CREATE TABLE "{table}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            {col_defs}
        )
        """
    )

    placeholders = ", ".join(["?"] * (3 + len(cols)))
    col_names = ", ".join([f'"{c}"' for c in cols])
    now = datetime.now(timezone.utc).isoformat()
    inserted = 0

    for row in rows[data_start_idx - 1 :]:
        r = list(row or [])
        if len(r) <= date_col_idx:
            continue
        if not is_date_like(r[date_col_idx]):
            continue
        sliced = r[date_col_idx : date_col_idx + len(cols)]
        if len(sliced) < len(cols):
            sliced += [None] * (len(cols) - len(sliced))
        payload = [path.name, sheet_name, now, *[to_cell_value(x) for x in sliced]]
        conn.execute(
            f'INSERT INTO "{table}" (source_file, sheet_name, imported_at, {col_names}) VALUES ({placeholders})',
            payload,
        )
        inserted += 1
    return table, inserted


def import_one_workbook(conn: sqlite3.Connection, path: Path) -> list[tuple[str, int]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    sheet_names = [s for s in wb.sheetnames if "disclaimer" not in s.lower()]
    wb.close()
    results: list[tuple[str, int]] = []
    for sheet in sheet_names:
        results.append(import_one_sheet(conn, path, sheet))
    return results


def main() -> None:
    missing = [str(p) for p in XLSX_FILES if not p.exists()]
    if missing:
        raise FileNotFoundError(f"Missing xlsx files: {missing}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    try:
        summary = []
        for f in XLSX_FILES:
            for table, count in import_one_workbook(conn, f):
                summary.append((f.name, table, count))
        conn.commit()
    finally:
        conn.close()

    print("Gold xlsx import complete (sheet-level):")
    for file_name, table, count in summary:
        print(f"- {file_name} -> {table} ({count} rows)")
    print(f"DB: {DB_PATH}")


if __name__ == "__main__":
    main()
